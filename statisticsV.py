#Обработчик
from openpyxl import load_workbook, Workbook
import statistics


wb = load_workbook(str(input('Введите название файла ')))   #Открытие выборки
sheet = wb.get_sheet_by_name('Лист1')
temp = []
temp2 = []
for row in sheet.rows:
	if type(row[0].value)!=str:
		temp.append(float(row[0].value))
		temp2.append(float(row[1].value))	



def statisticV(name):										#Расчет числовых характеристик
	print('Размер: ', len(name))
	print('Среднее: ', statistics.mean(name))
	print('Дисперсия: ', statistics.variance(name))
	print('СКО: ', statistics.stdev(name))
	try:
		print('Мода: ', statistics.mode(name))
	except:
		print('Мода: -')
	print('Медиана: ', statistics.median(name))
	
def sortedV(name, name2):											#Сортировка
	wb = Workbook()
	ws = wb.active
	ws.append(['Выборка 1', 'Выборка 2'])
	name = sorted(name)
	name2 = sorted(name2)
	for i in range(len(name)):
		ws.append([name[i], name2[i]])
	sortedName = str(input('Введите название файла (куда сохранить отсортированную выборку, без ".xlsx"): '))
	wb.save(sortedName + '.xlsx')


re = 1
while re<=3:
	re = int(input("""Действие:
	1. Вывести числовые характеристики 1 выборки
	2. Вывести числовые характеристики 2 выборки
	3. Отсортировать
	4. Выход
	Выбор: """))
	if re==1:
		statisticV(temp)
	if re==2:
		statisticV(temp2)
	if re==3:
		sortedV(temp, temp2)

		
		
	
	
