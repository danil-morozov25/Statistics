#Морозов Д., Григоренко Н., Цветков В.
#Критерий Манна-Уитни
#Используется для сравнения двух независимых выборок (могут быть разных размеров)
#Если какой-то элемент больше сравниваемого, то ему засчитывается 1 балл. Если элементы равны, то им засчитывается по 0,5 балла. Затем баллы элементов для каждой выборки суммируются, а меньшая полученная сумма является критерием – U-статистика. Если выборки не имеют существенных различий, то значение критерия должно быть больше критического значения для выборок соответствующего размера.


from scipy.stats import mannwhitneyu #добавление библиотеки для расчета по формуле
from openpyxl import Workbook, load_workbook

class Mannwhitneyu():  #создание класса
    def mann(name1, name2):  #добавление объектов
        stat, p = mannwhitneyu(name1, name2)
        alpha = 0.05  #табличный критерий
        if p > alpha:  #p - расчетный критерий
            return[stat, p, 'Распредления выборки равны (не отвергает H0)']  #вывод для пользователя
        else:
            return[stat, p, 'Распределения выборки не равны (отвергает H0)']
 #Проверка    
wb = load_workbook(str(input('Введите название файла ')))
sheet = wb['Манн-Уитни']
name1 = []  #выборка 1 (Exel)
name2 = []  #выборка 2 (Exel)
for row in sheet.iter_rows(min_row=5, min_col=2, max_col=2):
    for cell in row:
        if type(cell.value)!=str:
            name1.append(cell.value)

for row in sheet.iter_rows(min_row=5, min_col=3, max_col=3, max_row = 19):
    for cell in row:
        if type(cell.value)!=str:
            name2.append(cell.value)
a = Mannwhitneyu.mann(name1, name2)
print(a)
